import os
import fitz
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def get_excel_page_info(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Khong tim thay tep: {file_path}")

    doc = fitz.open(file_path)
    pages = []
    filename = os.path.basename(file_path)

    for i in range(len(doc)):
        page = doc.load_page(i)
        pages.append({
            "index": i,
            "page_num": i + 1,
            "width": float(page.rect.width),
            "height": float(page.rect.height),
            "file_path": file_path,
            "filename": filename,
        })

    doc.close()
    return pages


def get_excel_thumbnail(pdf_path, page_num, max_size=(120, 160)):
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(page_num)
        zoom_x = max_size[0] / page.rect.width
        zoom_y = max_size[1] / page.rect.height
        zoom = min(zoom_x, zoom_y, 1.0)
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return img
    finally:
        doc.close()


def get_page_excel_preview_text(file_path, page_index, max_chars=500):
    doc = fitz.open(file_path)
    try:
        page = doc.load_page(page_index)
        blocks = page.get_text("blocks")
        lines = []
        for b in blocks:
            if b[6] == 0:
                text = b[4].strip()
                if text:
                    lines.append((b[1], text))
        lines.sort(key=lambda x: x[0])
        result = "\n".join(t for _, t in lines)
        if len(result) > max_chars:
            result = result[:max_chars] + "..."
        return result if result else "(Khong co du lieu bang bieu)"
    finally:
        doc.close()


def _write_sheet_from_page(ws, page, page_label):
    blocks = page.get_text("blocks")
    blocks.sort(key=lambda b: (b[1], b[0]))

    rows_dict = {}
    for b in blocks:
        if b[6] == 0:
            y = round(b[1], 1)
            x = b[0]
            text = b[4].strip()
            if not text:
                continue
            found = False
            for key in rows_dict:
                if abs(key - y) < 8:
                    rows_dict[key].append((x, text))
                    found = True
                    break
            if not found:
                rows_dict[y] = [(x, text)]

    sorted_rows = sorted(rows_dict.items(), key=lambda item: item[0])
    sorted_rows_list = []
    for y_pos, cells in sorted_rows:
        cells.sort(key=lambda c: c[0])
        sorted_rows_list.append([c[1] for c in cells])

    if not sorted_rows_list:
        ws.cell(row=1, column=1, value="(Khong co du lieu)").font = Font(italic=True, color="999999")
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for r_idx, row_data in enumerate(sorted_rows_list):
        for c_idx, cell_text in enumerate(row_data):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=cell_text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            if r_idx == 0:
                cell.font = header_font
                cell.fill = header_fill

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 15, 15)
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def extract_to_excel(input_path, page_indices, output_dir, combine=False, progress_callback=None):
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Khong tim thay tep: {input_path}")

    doc = fitz.open(input_path)
    total = len(page_indices)

    if total == 0:
        doc.close()
        raise ValueError("Khong co trang nao de xuat.")

    base_name = os.path.splitext(os.path.basename(input_path))[0]
    output_files = []

    if combine:
        wb = Workbook()
        for i, page_idx in enumerate(page_indices):
            page = doc.load_page(page_idx)
            sheet_name = f"Trang_{page_idx + 1}"[:31]
            ws = wb.create_sheet(title=sheet_name) if i > 0 else wb.active
            ws.title = sheet_name
            _write_sheet_from_page(ws, page, f"Trang {page_idx + 1}")
            if progress_callback:
                progress_callback(int((i + 1) / total * 100))

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        output_path = os.path.join(output_dir, f"{base_name}.xlsx")
        wb.save(output_path)
        output_files.append(output_path)
    else:
        for i, page_idx in enumerate(page_indices):
            page = doc.load_page(page_idx)
            wb = Workbook()
            ws = wb.active
            ws.title = f"Trang_{page_idx + 1}"[:31]
            _write_sheet_from_page(ws, page, f"Trang {page_idx + 1}")
            output_path = os.path.join(output_dir, f"{base_name}_Trang_{page_idx + 1}.xlsx")
            wb.save(output_path)
            output_files.append(output_path)
            if progress_callback:
                progress_callback(int((i + 1) / total * 100))

    doc.close()
    return output_files
